"""
Actual data about Ukraine Covid-19 cases
Info was downloaded from https://app.powerbi.com/view?r=eyJrIjoiN2M1MTY1MDktZTY5Mi00OTE0LWFiMDAtMjM4NTY0YWU2MmI3IiwidCI6IjI4OGJmYmNmLTVhYjItNDk2MS04YTM5LTg2MDYxYWFhY2Q4NiIsImMiOjl9&amp;fbclid=IwAR3vOXvEK0l3SaGSAxZGgNyc4cNSi17wegJwcFX4oPefbVgUR16RsWDxGjg
"""

import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart.axis import DateAxis

data_area_type_hosp_dynamics = pd.read_csv('covid19_by_area_type_hosp_dynamics.csv')
data_actual_dynamics = pd.read_csv('covid19_by_settlement_actual.csv')
data_area_dynamics = pd.read_csv('covid19_by_settlement_dynamics.csv')
areas = data_area_type_hosp_dynamics['registration_area'].unique()
types = ['Підозри', 'Виявлено', 'Хворіє', 'Смертей', 'Госпіталізовано']
colors = ['red', 'orange', 'green', 'blue', 'black', 'purple']
lens = []
_dict = {'Підозри': 3, 'Виявлено': 4, 'Хворіє': 5, 'Смертей': 6, 'Одужало': 7, 'Госпіталізовано': 8}
analyzing_areas = pd.DataFrame()


def excel_helper(df):
    df = df.reset_index()
    wb = openpyxl.Workbook()
    sheet = wb.active
    for row in dataframe_to_rows(df, index=True, header=True):
        sheet.append(row)
    chart = create_chart('Миколаївська область')
    return sheet, chart, wb


def excel_ploting(df, save_to):
    sheet, chart, wb = excel_helper(df)
    data = Reference(sheet, min_col=3, min_row=1, max_col=8, max_row=len(df) + 2)
    dates = Reference(sheet, min_col=2, min_row=3, max_row=len(df) + 2)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(dates)
    sheet.add_chart(chart, 'K3')
    wb.save(save_to)


def create_chart(title):
    chart = LineChart()
    chart.height = 15
    chart.width = 25
    chart.title = title
    chart.y_axis.title = 'Frequency'
    chart.y_axis.crossAx = 500
    chart.x_axis = DateAxis(crossAx=100)
    chart.x_axis.number_format = 'yyyy-mm-dd'
    chart.x_axis.majorTimeUnit = "days"
    chart.x_axis.title = 'Date'
    return chart


def excel_ploting_advanced(df, save_to, _type, _all_areas, isAll):
    if isAll:
        _all_areas = areas

    wb = openpyxl.Workbook()
    sheet = wb.active

    for row in dataframe_to_rows(df.reset_index(), index=True, header=True):
        sheet.append(row)

    main_chart = create_chart('Області - ' + _type)

    main_chart.y_axis.crosses = "max"

    dates, data, end, start, index = [], [], 0, 2, 1
    for index, i in enumerate(lens):
        end = start + i

        sheet.insert_rows(end + 1)

        for row in sheet.iter_rows(min_row=start, max_row=start):
            row[_dict[_type] - 1].value = _all_areas[index]
        if index == 0:
            dates = Reference(sheet, min_col=2, min_row=start + 1, max_row=end)
        data.append(Reference(sheet, min_col=_dict[_type], min_row=start, max_row=end))
        start = end + 1

    for index, i in enumerate(data):
        sub_chart = create_chart('')
        sub_chart.add_data(i, titles_from_data=True)
        sub_chart.set_categories(dates)
        main_chart += sub_chart

    sheet.add_chart(main_chart, 'K2')
    wb.save(save_to)


# plot line graphics of data type
def ploting(df, area, *args):
    # multiple data type ploting
    for i in args:
        df.plot(kind='line', y=i, figsize=(12, 10), linewidth=4, ax=plt.gca(), linestyle="--")
    # set x-axis label
    plt.xlabel("Date", fontsize=18)
    # set y-axis label
    plt.ylabel('Count of patients', fontsize=21)
    # set title
    plt.title(area + " area - dynamics", fontsize=21)
    # grid on
    plt.grid(True)
    # ploting
    plt.show()


# function to read and parse data from file
def data_by_area(area):
    # define global variable
    global data_area_type_hosp_dynamics
    # store date about Covid-19 by given area
    total_by_area = data_area_type_hosp_dynamics.loc[data_area_type_hosp_dynamics['registration_area'] == area]
    # add column with area's name
    return total_by_area


# rename some columns
def rename(df):
    df.rename(columns={'total_susp': 'Підозр', 'total_confirm': 'Хворіє', 'total_death': 'Померло',
                       'total_recover': 'Одужало'}, inplace=True)


def parse_data_by_area(total_by_area):
    # group by column - 'zvit-date'
    total_by_area_grouped_date = total_by_area.groupby(['zvit_date']).sum()

    # find a number of patients who is under suspicion
    suspicion = total_by_area_grouped_date['new_susp'].cumsum()

    # find a number of detected cases
    detected = total_by_area_grouped_date['new_confirm'].cumsum()

    # find a number of lethal cases
    lethal = total_by_area_grouped_date['new_death'].cumsum()

    # find a number of cases of recovery
    recovered = total_by_area_grouped_date['new_recover'].cumsum()

    # find a number of patients who need to be hospitalized
    hospitalization = \
        total_by_area.loc[total_by_area['is_required_hospitalization'] == 'Так'].groupby('zvit_date').sum()[
            'new_susp'].cumsum()

    # rewrite columns data in dataset
    # rewrite suspicion column
    total_by_area_grouped_date['new_susp'] = suspicion

    # rewrite column of detected cases
    total_by_area_grouped_date['new_confirm'] = detected

    # rewrite column of lethal cases
    total_by_area_grouped_date['new_death'] = lethal

    # rewrite column of recovery cases
    total_by_area_grouped_date['new_recover'] = recovered

    # rewrite hospitalization column
    total_by_area_grouped_date['hospitalization'] = hospitalization

    # rename some columns
    total_by_area_grouped_date.rename(
        columns={'new_susp': 'Підозри', 'new_confirm': 'Виявлено', 'active_confirm': 'Хворіє',
                 'new_death': 'Смертей', 'new_recover': 'Одужало',
                 'hospitalization': 'Госпіталізовано'}, inplace=True)

    return total_by_area_grouped_date


def show_dynamics(df):
    # ploting dynamics in one figure
    ploting(df, 'Миколаївська', 'Підозри', 'Виявлено', 'Хворіє', 'Смертей', 'Одужало', 'Госпіталізовано')


def analyse_per_area():
    # define global variable with actual statistics
    global data_actual_dynamics

    # group by area and get the total number of sick patients
    data_per_area = data_actual_dynamics.groupby('registration_area')[['total_confirm']].sum()
    return data_per_area


def data_for_mapping():
    # define global variable with actual statistics one more time
    global data_actual_dynamics
    # array for coordinates
    lng, lat = [], []
    # loop for filling arrays
    for i in sorted(data_actual_dynamics['registration_region'].unique()):
        lng.append(data_actual_dynamics.loc[data_actual_dynamics['registration_region'] == i][
                       'registration_settlement_lng'].mean())
        lat.append(data_actual_dynamics.loc[data_actual_dynamics['registration_region'] == i][
                       'registration_settlement_lat'].mean())

    # get dataset for mapping by region
    total_sick_per_region = data_actual_dynamics.groupby(['registration_region'])[
        ['total_susp', 'total_confirm', 'total_death', 'total_recover']].sum()

    # rename some columns
    rename(total_sick_per_region)

    total_sick_per_area = data_actual_dynamics.groupby(['registration_area'])[
        ['total_susp', 'total_confirm', 'total_death', 'total_recover']].sum()

    # rename some columns
    rename(total_sick_per_area)

    # add this columns to dataset
    total_sick_per_region['lng'] = lng
    total_sick_per_region['lat'] = lat
    return total_sick_per_region, total_sick_per_area


def figure_sett(title):
    # figure title
    plt.title(title, fontsize=20)
    # x-axis label
    plt.xlabel('Date')
    # y-axis label
    plt.ylabel('Frequency')
    # grid on
    plt.grid()
    # showing
    plt.show()


def analyze_helper(i, type):
    global analyzing_areas
    # data by area
    total_by_area = data_by_area(i)
    # parsed data by area
    parsed = parse_data_by_area(total_by_area)
    parsed['Область'] = i
    lens.append(len(parsed))
    analyzing_areas = analyzing_areas.append(parsed)

    # ploting
    parsed.plot(kind='line', y=type, figsize=(12, 10), ax=plt.gca(), linewidth=4, linestyle="--", label=i)
    # add some settings


def analyze():
    all, type, analyze_areas = False, '', []
    # get option to analyze
    print('Analyze areas\n1 - all\n2 - selected areas')
    option = input()
    if option == '1':
        all = True
        print('Input type:')
        # available types
        for index, i in enumerate(types):
            print(index + 1, ' - ', i)
        # get type to plot
        type = types[int(input()) - 1]
        for i in areas:
            analyze_helper(i, type)
        # add some settings
        figure_sett(type)

    elif option == '2':
        # available areas
        for index, i in enumerate(areas):
            print(index + 1, ' - ', i)
        # get areas
        analyze_areas = [areas[i - 1] for i in list(map(int, input().split(' ')))]
        print('Input type:')
        for index, i in enumerate(types):
            print(index + 1, ' - ', i)
        # get type to plot
        type = types[int(input()) - 1]
        for i in analyze_areas:
            analyze_helper(i, type)
        # add some settings
        figure_sett(type)
    return all, type, analyze_areas


def main():
    # printing all columns
    pd.set_option('display.max_columns', None)
    # printing all rows
    pd.set_option('display.max_rows', None)

    # data by Mykolayivska area
    df = data_by_area('Миколаївська')
    # grouped data by Mykolayivska area
    total_by_area_grouped = parse_data_by_area(df)

    # show graphics
    show_dynamics(total_by_area_grouped)

    # analyze per area
    isAll, type, all_areas = analyze()

    # dataset to plot in a map
    mapping_region, mapping_area = data_for_mapping()

    # save to cvs files
    mapping_region.to_csv('mapping_per_region.csv')
    mapping_area.to_csv('mapping_per_area.csv')

    # printing
    print('Grouped date by Mykolayivska area\n', total_by_area_grouped)
    print('\nDate for mapping by region\n', mapping_region)
    print('\nDate for mapping by area\n', mapping_area)
    print('\nAnalyzed date per area\n', analyzing_areas)
    print(analyzing_areas.loc[analyzing_areas['Область'] == 'Вінницька'])
    # save to excel and plot
    excel_ploting(total_by_area_grouped, 'total_by_area_grouped.xlsx')
    excel_ploting_advanced(analyzing_areas, 'analyzing_areas.xlsx', type, all_areas, isAll)

    mapping_region.to_excel('mapping_region.xlsx')
    mapping_area.to_excel('mapping_area.xlsx')


# run the program
if __name__ == '__main__':
    main()
